"""
Grade export endpoints — returns .xlsx or .pdf files.
"""

import io
from datetime import datetime

from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status

from apps.users.permissions import IsProfessor, IsAdmin, IsAdminOrProfessor
from apps.pedagogique.models import Major
from apps.assignments.models import Submission


def _get_grades_data(major_id):
    """Build a list of grade rows for a given major."""
    try:
        major = Major.objects.get(pk=major_id)
    except Major.DoesNotExist:
        return None, None

    submissions = Submission.objects.filter(
        assignment__majors=major,
        student__isnull=False,
    ).select_related(
        'student__user', 'assignment'
    ).order_by('student__user__last_name', 'assignment__title')

    rows = []
    for sub in submissions:
        rows.append({
            'student': sub.student.user.get_full_name(),
            'assignment': sub.assignment.title,
            'type': sub.assignment.type,
            'grade': str(sub.grade) if sub.grade is not None else 'N/A',
            'status': sub.status,
            'submitted_at': sub.submitted_at.strftime('%Y-%m-%d %H:%M') if sub.submitted_at else '',
        })
    return major, rows


@api_view(['GET'])
@permission_classes([IsAdminOrProfessor])
def export_grades(request):
    major_id = request.query_params.get('major')
    fmt = request.query_params.get('format', 'xlsx')

    if not major_id:
        return Response({'detail': 'major query param is required.'}, status=status.HTTP_400_BAD_REQUEST)

    major, rows = _get_grades_data(major_id)
    if major is None:
        return Response({'detail': 'Major not found.'}, status=status.HTTP_404_NOT_FOUND)

    if fmt == 'xlsx':
        return _export_xlsx(major, rows)
    elif fmt == 'pdf':
        return _export_pdf(major, rows)
    else:
        return Response({'detail': 'format must be xlsx or pdf.'}, status=status.HTTP_400_BAD_REQUEST)


def _export_xlsx(major, rows):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Grades — {major.name}'

    headers = ['Student Name', 'Assignment', 'Type', 'Grade (/20)', 'Status', 'Submitted At']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1E293B', end_color='1E293B', fill_type='solid')

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row_idx, row in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=row['student'])
        ws.cell(row=row_idx, column=2, value=row['assignment'])
        ws.cell(row=row_idx, column=3, value=row['type'])
        ws.cell(row=row_idx, column=4, value=row['grade'])
        ws.cell(row=row_idx, column=5, value=row['status'])
        ws.cell(row=row_idx, column=6, value=row['submitted_at'])

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f'grades_{major.code}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _export_pdf(major, rows):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    styles = getSampleStyleSheet()
    elements = []

    # Title
    title = Paragraph(
        f'<b>Grade Report — {major.name} ({major.code})</b>',
        styles['Title'],
    )
    elements.append(title)
    elements.append(Spacer(1, 0.5 * cm))
    elements.append(Paragraph(f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', styles['Normal']))
    elements.append(Spacer(1, 0.5 * cm))

    headers = ['Student Name', 'Assignment', 'Type', 'Grade (/20)', 'Status', 'Submitted At']
    table_data = [headers] + [
        [r['student'], r['assignment'], r['type'], r['grade'], r['status'], r['submitted_at']]
        for r in rows
    ]

    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F5F9')]),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#CBD5E1')),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
    ]))
    elements.append(table)
    doc.build(elements)

    buf.seek(0)
    filename = f'grades_{major.code}_{datetime.now().strftime("%Y%m%d")}.pdf'
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
